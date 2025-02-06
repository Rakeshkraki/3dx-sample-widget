import openpyxl
import re

excel_file_path = 'C:\\Users\\rakesh.k\\HomeScreenShortcuts\\createAttributeExel_rev_3\\CreateAttribute_rev3.xlsx'
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

with open('C:\\Users\\rakesh.k\\HomeScreenShortcuts\\createAttributeExel_rev_3\\attribute.tcl', 'w') as txt_file:
    num_columns = sheet.max_column

    # Iterate through each column
    for col in range(1, num_columns + 1):

        first_cell_value = str(sheet.cell(row=1, column=col).value).strip()
        first_cell_value = re.sub(r'[^A-Za-z0-9()\[\]]', '', first_cell_value)
        second_cell_value = str(sheet.cell(row=2, column=col).value).strip()
        third_cell_value = str(sheet.cell(row=3, column=col).value)

        if second_cell_value == 'Drop down/String':
            range_values1 = []
            for row in range(3, sheet.max_row + 1):
                third_cell_value = (sheet.cell(row=row, column=col).value)
                if third_cell_value is None:
                    break
                range_values1.append(str(third_cell_value))

            if range_values1:
                range_str = " ".join([f"range = '{val.strip()}'" for val in range_values1])
                txt_file.write(f"add attribute 'tmeic_{first_cell_value}' type String {range_str} property application value \"\" property version value 'V6R2023x' property installer value \"\" property 'installed date' value \"\" property 'original name' value 'tmeic_{first_cell_value}' property classificationAttributes value true;\n\n")

        elif second_cell_value == 'Drop down/Integer':
            range_values2 = []
            for row in range(3, sheet.max_row + 1):
                third_cell_value = sheet.cell(row=row, column=col).value
                if third_cell_value is None:
                    break
                range_values2.append(str(third_cell_value))

            if range_values2:
                range_str = " ".join([f"range = {val}" for val in range_values2])
                txt_file.write(f"add attribute 'tmeic_{first_cell_value}' type Integer {range_str} property application value \"\" property version value 'V6R2023x' property installer value \"\" property 'installed date' value \"\" property 'original name' value 'tmeic_{first_cell_value}' property classificationAttributes value true;\n\n")

        elif first_cell_value != "None":
            txt_file.write(f"add attribute 'tmeic_{first_cell_value}' type {second_cell_value} property application value \"\" property version value 'V6R2023x' property installer value \"\" property 'installed date' value \"\" property 'original name' value 'tmeic_{first_cell_value}' property classificationAttributes value true;\n\n")

print("Data from all columns has been written to new.txt")
